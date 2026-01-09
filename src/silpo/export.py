import csv
import os
from datetime import datetime
from typing import List, Tuple, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 4000) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 60)

def export_csv(exports_dir: str, header: List[str], rows: List[Tuple[Any, ...]]) -> str:
    os.makedirs(exports_dir, exist_ok=True)
    out = os.path.join(exports_dir, "latest.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    return out

def export_xlsx(exports_dir: str, header: List[str], rows: List[Tuple[Any, ...]]) -> str:
    os.makedirs(exports_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "observations"
    ws.append(header)
    for r in rows:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
    _autosize(ws)
    out = os.path.join(exports_dir, "latest.xlsx")
    wb.save(out)

    # versioned copy
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    ver = os.path.join(exports_dir, f"silpo_{ts}.xlsx")
    wb.save(ver)
    return out
