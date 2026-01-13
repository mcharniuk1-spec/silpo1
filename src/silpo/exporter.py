import csv
import os
from datetime import datetime
from typing import List, Tuple
import sqlite3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row, 5000) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)

def export_xlsx_csv(
    conn: sqlite3.Connection,
    exports_dir: str,
    run_id: str,
    log_events: List[dict],
) -> Tuple[str, str]:
    os.makedirs(exports_dir, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    # Fetch data
    products = conn.execute(
        """
        SELECT upload_ts, page_number, page_url, source, product_id, product_url, title, brand,
               pack_qty, pack_unit, price_current, price_old, discount_pct
        FROM products
        WHERE run_id=?
        ORDER BY page_number, title
        """,
        (run_id,),
    ).fetchall()

    page_logs = conn.execute(
        """
        SELECT upload_ts, page_number, page_url, method, status, http_status, items_seen, items_saved, note
        FROM page_logs
        WHERE run_id=?
        ORDER BY page_number
        """,
        (run_id,),
    ).fetchall()

    # XLSX
    wb = Workbook()

    ws = wb.active
    ws.title = "products"
    prod_header = ["upload_ts","page_number","page_url","source","product_id","product_url","title","brand","pack_qty","pack_unit","price_current","price_old","discount_pct"]
    ws.append(prod_header)
    for r in products:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    _autosize(ws)

    ws2 = wb.create_sheet("page_logs")
    pl_header = ["upload_ts","page_number","page_url","method","status","http_status","items_seen","items_saved","note"]
    ws2.append(pl_header)
    for r in page_logs:
        ws2.append(list(r))
    ws2.freeze_panes = "A2"
    _autosize(ws2)

    ws3 = wb.create_sheet("logs")
    ws3.append(["ts","level","event","message"])
    for e in log_events:
        ws3.append([e.get("ts",""), e.get("level",""), e.get("event",""), e.get("message","")])
    ws3.freeze_panes = "A2"
    _autosize(ws3)

    xlsx_path = os.path.join(exports_dir, f"silpo_{ts}_{run_id[:8]}.xlsx")
    latest_xlsx = os.path.join(exports_dir, "latest.xlsx")
    wb.save(xlsx_path)
    wb.save(latest_xlsx)

    # CSV (products only)
    csv_path = os.path.join(exports_dir, f"silpo_{ts}_{run_id[:8]}.csv")
    latest_csv = os.path.join(exports_dir, "latest.csv")
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(prod_header)
        for r in products:
            w.writerow(list(r))
    with open(latest_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(prod_header)
        for r in products:
            w.writerow(list(r))

    # Hard guarantees (if these fail — run must be red)
    for p in (xlsx_path, latest_xlsx, csv_path, latest_csv):
        if not os.path.exists(p):
            raise RuntimeError(f"Export file not created: {p}")

    return latest_xlsx, latest_csv
